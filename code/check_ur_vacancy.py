#!/usr/bin/env python3
"""
UR団地空室チェッカー
CSVファイルの団地リストを読み込み、2DK/2LDK/3DK/3LDKの空室を確認してCSV/Excelに出力する
"""

import csv
import json
import re
import time
import urllib.request
import urllib.parse
from datetime import datetime
from pathlib import Path

# ---- 設定 ----
CSV_INPUT   = Path(__file__).parent.parent / "ur_agent" / "ur_danchi.csv"
OUTPUT_DIR  = Path(__file__).parent.parent / "ur_agent" / "aval"
TARGET_TYPES = {"2DK", "2LDK", "3DK", "3LDK"}
API_URL     = "https://chintai.r6.ur-net.go.jp/chintai/api/bukken/detail/detail_bukken_room/"
REQUEST_INTERVAL = 1.0  # サーバー負荷軽減のため各リクエスト間に待機(秒)


def extract_shisya_danchi(ur_url: str) -> tuple[str, str] | None:
    """
    URページURLからshisya/danchiパラメータを抽出する
    例: .../40_2690.html  →  shisya='40', danchi='269'
    パターン: XX_YYYY.html → shisya=XX, danchi=str(int(YYYY/10))
    """
    m = re.search(r'/(\d+)_(\d+)\.html', ur_url)
    if not m:
        return None
    shisya = m.group(1)
    raw    = m.group(2)          # e.g. "2690"
    danchi = str(int(raw) // 10) # "2690" → 269, "4020" → 402
    return shisya, danchi


def fetch_rooms(shisya: str, danchi: str, shikibetu: str = "0") -> list[dict]:
    """全ページを取得して全空室リストを返す"""
    all_rooms = []
    page = 0
    page_max = 1

    while page <= page_max:
        data = urllib.parse.urlencode({
            "shisya":       shisya,
            "danchi":       danchi,
            "shikibetu":    shikibetu,
            "orderByField": "0",
            "orderBySort":  "0",
            "pageIndex":    str(page),
            "sp":           "",
        }).encode()

        req = urllib.request.Request(
            API_URL,
            data=data,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Referer":      "https://www.ur-net.go.jp/",
                "User-Agent":   "Mozilla/5.0",
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                body = resp.read().decode("utf-8")
        except Exception as e:
            print(f"  [ERROR] shisya={shisya} danchi={danchi} page={page}: {e}")
            break

        if body.strip() == "null":
            break

        rooms = json.loads(body)
        if not rooms:
            break

        # 1件目にページ情報が入っている
        page_max = int(rooms[0].get("pageMax", 1)) - 1
        all_rooms.extend(rooms)
        page += 1

        if page <= page_max:
            time.sleep(REQUEST_INTERVAL)

    return all_rooms


def check_vacancies() -> list[dict]:
    """CSVを読み込み、各団地の対象間取り空室をチェックして結果リストを返す"""
    results = []

    with open(CSV_INPUT, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    total = len(rows)
    for i, row in enumerate(rows, 1):
        name    = row["団地名"]
        ur_url  = row["団地URページ"]
        station = row["最寄り駅（二子玉川から）"]
        travel  = row["二子玉川からの所要時間"]

        print(f"[{i}/{total}] {name} を確認中...")

        params = extract_shisya_danchi(ur_url)
        if params is None:
            print(f"  URLパターン不明: {ur_url}")
            continue

        shisya, danchi = params
        rooms = fetch_rooms(shisya, danchi)

        matched = [r for r in rooms if r.get("type", "") in TARGET_TYPES]

        if not matched:
            print(f"  → 対象間取りの空室なし")
            results.append({
                "番号":                 row["番号"],
                "団地名":              name,
                "最寄り駅":            station,
                "二子玉川からの所要時間": travel,
                "間取り":              "なし",
                "部屋名":              "-",
                "月額家賃":            "-",
                "共益費":              "-",
                "床面積":              "-",
                "階":                  "-",
                "URページ":            ur_url,
            })
        else:
            print(f"  → {len(matched)}件 ヒット ({[r['type'] for r in matched]})")
            for r in matched:
                results.append({
                    "番号":                 row["番号"],
                    "団地名":              name,
                    "最寄り駅":            station,
                    "二子玉川からの所要時間": travel,
                    "間取り":              r.get("type", ""),
                    "部屋名":              r.get("name", ""),
                    "月額家賃":            r.get("rent", ""),
                    "共益費":              r.get("commonfee", ""),
                    "床面積":              re.sub(r'&#\d+;', '㎡', r.get("floorspace", "")),
                    "階":                  r.get("floor", ""),
                    "URページ":            ur_url,
                })

        time.sleep(REQUEST_INTERVAL)

    return results


def save_csv(results: list[dict], path: Path):
    if not results:
        return
    fieldnames = list(results[0].keys())
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)
    print(f"CSV保存: {path}")


def save_excel(results: list[dict], path: Path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl未インストールのためExcel出力をスキップします")
        return

    if not results:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UR空室チェック"

    headers = list(results[0].keys())

    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill(fill_type="solid", fgColor="2E5FA3")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    d_font = Font(name="Calibri", size=10)
    d_align = Alignment(vertical="center", wrap_text=True)
    alt_fill = PatternFill(fill_type="solid", fgColor="EBF1F8")
    no_room_fill = PatternFill(fill_type="solid", fgColor="F5F5F5")
    hit_fill = PatternFill(fill_type="solid", fgColor="E8F5E9")

    # ヘッダー行
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font   = h_font
        cell.fill   = h_fill
        cell.alignment = h_align
        cell.border = border
    ws.row_dimensions[1].height = 32

    # データ行
    for r, row in enumerate(results, 2):
        is_hit = row["間取り"] != "なし"
        for c, key in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=row[key])
            cell.font   = d_font
            cell.alignment = d_align
            cell.border = border
            if is_hit:
                cell.fill = hit_fill
            elif r % 2 == 0:
                cell.fill = alt_fill
            else:
                cell.fill = no_room_fill
        ws.row_dimensions[r].height = 22

    # 列幅
    col_widths = {1:6, 2:22, 3:16, 4:14, 5:8, 6:16, 7:12, 8:10, 9:8, 10:6, 11:55}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # URページ列をハイパーリンク
    url_col = headers.index("URページ") + 1
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=url_col)
        if cell.value and str(cell.value).startswith("http"):
            cell.hyperlink = cell.value
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"Excel保存: {path}")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"=== UR空室チェック開始: {ts} ===")
    print(f"対象間取り: {', '.join(sorted(TARGET_TYPES))}")
    print(f"入力CSV: {CSV_INPUT}\n")

    results = check_vacancies()

    csv_path   = OUTPUT_DIR / f"vacancy_{ts}.csv"
    excel_path = OUTPUT_DIR / f"vacancy_{ts}.xlsx"

    save_csv(results, csv_path)
    save_excel(results, excel_path)

    hit_count = sum(1 for r in results if r["間取り"] != "なし")
    print(f"\n=== 完了 ===")
    print(f"総団地数: {len(set(r['団地名'] for r in results))}")
    print(f"空室あり: {hit_count}件")
    print(f"出力先: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
